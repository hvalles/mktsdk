import os
from  enum import Enum
from openpyxl import load_workbook
from model import Producto, Categoria, Componente, Kit, Variacion, Stock, Precio, Model

class ExcelType(Enum):
    ## Template clasification
    MASTER = 1
    KITS = 2
    PRICE = 4
    STOCK = 8
    VARIACION = 16

class Excelfile():
    COL_CATEGORIA=10
    def __init__(self, path, template:ExcelType):
        self.fname =path
        self.template = template
        self.first_row = 2
        self.wb = None
        self.hoja = None
        self.last_row = 1
        self.last_col = 1
        self.first_row = 1
        self.header = 1
        self.category:Categoria=None
        self.skus=[]
        self.master=0
   
    def start_row(self):
        "According to template return what row start data to consume"
        if self.template in [ExcelType.MASTER, ExcelType.STOCK]: return 3
        return 2

    def new_model(self):
        "Creates  a new object according to template"
        if self.template==ExcelType.MASTER: return Producto()
        if self.template==ExcelType.KITS: return Kit()
        if self.template==ExcelType.PRICE: return Precio()
        if self.template==ExcelType.STOCK: return Stock()
        if self.template==ExcelType.VARIACION : return Variacion()
        raise TypeError()

    def parse_kit(self, row:int=0, previo=None):
        # Handled by positional columns
        sku = self.hoja.cell(row, 1).value
        if not previo or previo.sku!=sku:
            r = Kit()
            r.sku = sku
            r.comentario = self.hoja.cell(row, 2).value
        c:Componente=Componente()
        c.sku=self.hoja.cell(row, 3).value
        c.cantidad=self.hoja.cell(row, 4).value
        r.componentes.append(c)
        return r

    def parse_model(self, row:int=0, primary=True):
        "Read Excel row and return model object"
        r:Model = None
        if row:
            if self.template==ExcelType.MASTER:
                r = Producto() if primary else Variacion()
            else:
                r = self.new_model()
                   
            # Named columns
            for i in range(1,self.last_col+1):                
                col = self.hoja.cell(self.header, i).value
                if not col: continue
                value = self.hoja.cell(row, i).value
                z=r.haskey(col)
                if z:
                    if not value: value=''
                    r[col] = value
                if i>self.master and self.category and value and r.haskey('atributos') \
                    and len(self.category._mapa)>=(i-(self.master+1)):
                    col = self.category._mapa[i - (self.master+1)]
                    if primary and self.category.isnotVariant(col):
                        r.atributos.append({"atributo":col, 
                            "valor":self.hoja.cell(row, i).value})
                    if not primary and self.category.isVariant(col):
                        r.atributos.append({"atributo":col, 
                            "valor":self.hoja.cell(row, i).value})

        return r

    def read_row(self):
        "Generator for each row into model"
        previo = None
        for i in range(self.first_row, self.last_row+1):
            if not self.hoja[f"A{i}"].value: continue
            if self.template==ExcelType.MASTER: # Yields 2 records
                sku = self.hoja[f"A{i}"].value
                if not sku in self.skus:
                    yield self.parse_model(i)
                    self.skus.append(sku)
                yield self.parse_model(i, False)
            else:
                if self.template==ExcelType.KITS: # Yields by group
                    r = self.parse_kit(i, previo)
                    if previo is None: previo = r
                    if previo.sku!=r.sku:
                        yield previo
                        previo = r
                else:
                    yield self.parse_model(i)
        if previo: yield previo

    def open(self, sheet=None):
        "Open excel file and set initial values"
        if not os.path.exists(self.fname): raise FileExistsError()
        self.wb = load_workbook(filename=self.fname, data_only=True)
        self.hoja = self.wb.active
        if sheet: self.hoja = self.wb[sheet]
        self.last_row = self.hoja.max_row
        self.last_col = self.hoja.max_column
        self.first_row = self.start_row()
        self.header = self.first_row - 1
        return self

